import os
import pickle
import openpyxl
import traceback
from lib import eda_parameters as eda
from lib import digikey_api as dk

# Update this to the board you're working on
this_sheet_name = "Status LED Ring"
pickle_path = os.path.join('digikey_cache', 'digikey_data.pickle')
dk.setup()

def save_pickle(dictionary_list):
    with open(pickle_path, 'wb') as f:
        pickle.dump(dictionary_list, f)
    print("Saved Digikey data to pickle")


file_found = False
for file in os.listdir(os.getcwd()):
    if file.startswith("altium_helper"):
        continue
    if file.endswith(".xlsx"):
        file_found = True
        wb = openpyxl.load_workbook(file)
        print("Loaded file: " + file)
        break
if not file_found:
    print("Error: No .xlsx file found")
    exit(1)

# Check if the output sheet exists
if "Output" not in wb.sheetnames:
    print("Error: Output sheet not found")
    exit(1)

# Check if the board sheet exists
if this_sheet_name not in wb.sheetnames:
    print("Error: Board sheet not found")
    exit(1)

output = wb["Output"]  # Copy vals from Altium into this sheet
board = wb[this_sheet_name]

alitum_params = eda.sheet_to_param_list(output, eda.AltiumParameter)
kicad_params = eda.sheet_to_param_list(board, eda.KiCadParameter, True)

combined = wb["Combined"]  # This is an aggregate of all internal part numbers
combined_mfns = []
combined_internal_pns = []
for row in combined.iter_rows(min_row=2):
    combined_mfns.append(str(row[3].value))
    internal_1 = str(row[5].value)
    internal_2 = str(row[6].value)
    if internal_1 is None or internal_1 == 'None':
        combined_internal_pns.append(internal_2)
    else:
        combined_internal_pns.append(internal_1)

dictionary_list = []
has_digikey_data = False
# If the Digikey data has already been downloaded, load it from the pickle
if os.path.isfile(pickle_path):
    has_digikey_data = True
    with open(pickle_path, 'rb') as f:
        dictionary_list = pickle.load(f)

# Update the Digikey data for each part
try:
    for altium_param in alitum_params:
        skip = False
        if altium_param.notes == 'DNP':
            skip = True
        for ref in ['FID', 'LOGO', 'TP', 'MP', 'H', 'JP']:
            if ref in altium_param.identifier:
                skip = True
                continue
        if skip:
            print("Skipping " + altium_param.identifier)
            continue

        mfn = ''
        for kicad_param in kicad_params:
            if (altium_param.identifier + ',') in kicad_param.references:
                mfn = kicad_param.manufacturer_pn
        if mfn == '':  # Try again without the comma
            for kicad_param in kicad_params:
                if altium_param.identifier in kicad_param.references:
                    mfn = kicad_param.manufacturer_pn
        if mfn == '':
            print("No BOM item for " + altium_param.identifier)
            continue

        print("Looking up " + mfn)
        product = None
        if has_digikey_data:
            for dictionary in dictionary_list:
                if mfn in str(dictionary):
                    print("Found in pickle")
                    product = dictionary
                    break
        if product is None:
            product = dk.get_product(mfn)
            original_mfn = mfn
            while product['exact_manufacturer_products_count'] == 0:
                print("Could not find " + mfn)
                mfn = input("Enter a new mfn: ")
                if len(mfn) == 0 or len(mfn) > 100:
                    print("Invalid mfn")
                    continue
                # Update the mfn in the kicad_params list
                for param in kicad_params:
                    if param.manufacturer_pn == original_mfn:
                        param.manufacturer_pn = mfn
                # Update the mfn in the combined list
                combined_mfns[combined_mfns.index(original_mfn)] = mfn
                # Get the new product
                product = dk.get_product(mfn)
            print("Found on Digikey")
            dictionary_list.append(product)

        # Update parameters
        exact_mfg_product = product['exact_manufacturer_products'][0]
        altium_param.cost_1 = dk.get_cost_1(exact_mfg_product)
        altium_param.cost_1000 = dk.get_cost_1000(product)
        if altium_param.current != '*':
            for param in exact_mfg_product['parameters']:
                if 'Current' in param['parameter']:
                    altium_param.current = param['value']
        altium_param.distributor = 'Digikey'
        altium_param.distributor_pn = str(
            exact_mfg_product['digi_key_part_number'])
        altium_param.help_url = exact_mfg_product['primary_datasheet']
        altium_param.manufacturer = exact_mfg_product['manufacturer']['value']
        altium_param.manufacturer_pn = str(
            exact_mfg_product['manufacturer_part_number'])
        altium_param.internal_pn = combined_internal_pns[combined_mfns.index(
            mfn)]
        if altium_param.rohs != '*':
            altium_param.rohs = 'Yes' if 'ROHS3 Compliant' in str(
                product) else 'No'
        if altium_param.tolerance != '*':
            found = False
            for param in exact_mfg_product['parameters']:
                if 'Tolerance' in param['parameter']:
                    found = True
                    altium_param.tolerance = param['value']
            if not found:
                altium_param.tolerance = 'n/a'
        if altium_param.value != '*':
            full_value = exact_mfg_product['parameters'][-1]['value']
            full_value = full_value.replace(' µF', 'µF').replace(' pF', 'pF').replace(
                ' A', 'A').replace(' mH', 'mH').replace(' uH', 'uH').replace(' kOhms', 'k')
            # Discard ohms units for resistors
            trimmed_value = full_value.split(' ')[0]
            # Make 10000pF -> 10nF
            if trimmed_value == '10000pF':
                trimmed_value = '10nF'
            altium_param.value = trimmed_value
        if altium_param.voltage != '*':
            for param in exact_mfg_product['parameters']:
                if 'Voltage' in param['parameter']:
                    altium_param.voltage = param['value']
        if altium_param.wattage != '*':
            for param in exact_mfg_product['parameters']:
                if 'Power' in param['parameter']:
                    full_value = param['value']
                    # Prefer fractional wattage
                    if ', ' in full_value:
                        full_value = full_value.split(', ')[1]
                    altium_param.wattage = full_value
except:
    print("Error: Could not update parameters")
    print(traceback.format_exc())
    save_pickle(dictionary_list)

for altium_param in alitum_params:
    try:
        # Easy way to check if a part was found, but not perfectly
        # There will probably be more of these
        if '$' in altium_param.cost_1 and altium_param.help_url == '':
            print("Manual check required for " + altium_param.identifier)
            continue
        else:
            altium_param.fill_empty()
    except:
        print("Manual check required for " + altium_param.identifier)
        print(traceback.format_exc())

# Write the updated parameters back to the output sheet
eda.param_list_to_sheet(alitum_params, output)

# Save the workbook as a new file
wb.save("altium_helper.xlsx")

# Save the Digikey data as a pickle to save on API calls
save_pickle(dictionary_list)
